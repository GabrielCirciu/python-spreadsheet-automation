import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def discount_calculator():
    file_name = input('Current spreadsheet name: ')
    workbook = xl.load_workbook(f'{file_name}.xlsx')
    sheet_name = input('Current sheet name: ')
    sheet = workbook[sheet_name]
    sheet.cell(1, 4).value = 'discounted'

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    chart_values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(chart_values)
    sheet.add_chart(chart, 'f2')

    workbook.save(f'{file_name}.xlsx')


discount_calculator()
